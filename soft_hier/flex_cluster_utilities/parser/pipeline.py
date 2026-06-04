#!/usr/bin/env python3
"""
Compute L1 transformer performance tables and (optionally) generate a GVSOC
pipeline test from a START/END-delimited template.

This single script merges the former compute_l1.py + pipeline_ctt.py +
pipeline_cwt.py.  Running it prints the performance tables and generates the
per-pool GVSOC test directories.  With no --mode it prints both cluster splits
(serialized + overlapped) and generates all three flavours; with
--mode {seq,ctt,cwt} it prints only that mode's split and generates only its
test directory:

    * seq -- sequential pipeline: one stage per cluster, gated by `if (cid == i)`
      + a barrier after each, so stages run one after another and a single item
      flows through (the old compute_l1.py --gen-tests).
    * ctt -- compute-then-transfer streaming pipeline: --iterations N items are
      streamed as a wavefront (FILL -> STEADY -> DRAIN), single L1 buffer per
      cluster, compute and DMA SERIALIZED.
    * cwt -- compute-while-transfer streaming pipeline: same wavefront but with a
      DOUBLE-buffered L1 (2x footprint) and a core-role split so compute and DMA
      OVERLAP.

For ctt/cwt each stage emits its exact ordered sequence of layers (one
`mempool_wait()` placeholder per layer, block reps in a `for` loop); the seq test
spins once per cluster on the cluster's aggregate latency.  The simulation is
timing-only (placeholder data, mempool_wait spins).

By default no preload is generated.  With --gen-preload the HBM preload .elf is
built directly here (requires the riscv32-unknown-elf-* toolchain on PATH).

The template (xlsx, ods or csv) is parsed without markers: the header row (the
one labelling N_repetitions / FLOPs / Weights / Data_in / Data_out) is located
automatically.  Layers are read from the row below it until the trailing
parameter block -- reading stops at the first row whose block-name cell
(column A) is filled but whose operation cell (column C) is empty, since a real
block header always carries its first layer's operation on the same row.
Anything above the header or below that stop point (parameter tables, notes or
scratch totals) is ignored, so those can live at the top or bottom of the file.

Column mapping in the data region:
    * Block name      -> column A          (no header expected)
    * Operation name  -> column C          (no header expected)
    * GEMM marker     -> column X          (no header expected; "GEMM" string)
    * N_repetitions   -> located by header
    * FLOPs           -> located by header
    * Weights         -> located by header
    * Data_in         -> located by header
    * Data_out        -> located by header
Weights / Data_in / Data_out are read in elements; bytes are computed as
2 x elements (the spreadsheet uses FP16).

Hardware peak performance and utilisation are read from a JSON config:
    {
        "PE_peak":        {"Terapool": 3.7, "TensorPool": 1.0},
        "GEMM_peak":      {"Terapool": 1.1, "TensorPool": 6.62},
        "PE_utilization": 0.5,
        "throughput":     10.0,
        "bandwidth_gbs":  64
    }
PE_peak and GEMM_peak are in TFLOPS. PE_utilization is a fraction in [0, 1]
applied only to PE_peak (non-GEMM ops). throughput is the per-cluster
wall-clock budget in milliseconds, interpreted as compute + transfer time
(optional; omit to skip cluster splitting). bandwidth_gbs is the
inter-cluster/HBM bandwidth in GB/s used to estimate transfer time as
bytes/(bandwidth_gbs*1e6) ms (optional; omit or 0 to count compute only).
Each cluster's transfer is its last layer's data_out hand-off, plus the HBM
data_in load on cluster 0.

The script reproduces the three reference tables:
    * K1:M3   -> FLOPs/ms per pool, split non-GEMM / GEMM
    * T52:W54 -> L1 footprint (compute-then-transfer vs compute-while-transfer)
    * Y5:Z53  -> Per-row, per-block and total latency [ms] per pool

For xlsx/ods, the script reads the *cached* calculated values of formula cells.
If a spreadsheet was edited but never opened in a calculator (Excel /
LibreOffice), reopen and save it once so the cached values are refreshed.

Usage:
    python pipeline.py final.xlsx config.json                       # both splits, all 3 modes
    python pipeline.py final.csv  config.json --mode ctt --iterations 8
    python pipeline.py final.csv  config.json --mode cwt --iterations 8 --gen-preload
    python pipeline.py final.csv  config.json --mode seq
    python pipeline.py final.csv  config.json --plot                # also save split bar plots
"""

import argparse
import csv
import json
import math
import os
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Logical field -> accepted header names in the header row
HEADER_ALIASES: Dict[str, List[str]] = {
    'n_rep':   ['N_repetitions', 'N_rep', 'Repetitions'],
    'flops':   ['FLOPs', 'Flops', 'FLOPS'],
    'weights': ['Weights'],
    'din':     ['Data_in'],
    'dout':    ['Data_out'],
}

# Fallback positions (0-based) for columns that don't carry a header label
DEFAULT_COLS: Dict[str, int] = {
    'block': 0,   # column A
    'op':    2,   # column C
    'gemm':  23,  # column X
}

BYTES_PER_ELEM = 2  # spreadsheet uses *2 (FP16)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class LayerRow:
    excel_row:     int
    block:         str
    n_rep:         int
    op:            str
    flops:         float
    weights_bytes: float
    din_bytes:     float
    dout_bytes:    float
    is_gemm:       bool


@dataclass
class Block:
    name:  str
    n_rep: int
    rows:  List[LayerRow] = field(default_factory=list)


@dataclass
class ClusterSegment:
    """A contiguous slice of repetitions of one block assigned to a cluster."""
    block:    Block
    rep_from: int   # 0-based, inclusive
    rep_to:   int   # 0-based, inclusive

    @property
    def n_reps(self) -> int:
        return self.rep_to - self.rep_from + 1


@dataclass
class ClusterSplit:
    idx:                       int
    segments:                  List[ClusterSegment]
    latency:                   float   # total compute latency [ms] for this cluster
    transfer_ms:               float   # hand-off transfer time [ms] (dout/BW, +din/BW for cluster 0)
    first_weights:             float   # weights_bytes of the cluster's first layer
    first_din:                 float   # din_bytes of the cluster's first layer
    last_weights:              float   # weights_bytes of the cluster's last layer
    last_dout:                 float   # dout_bytes of the cluster's last layer (pipeline output)
    l1_compute_then_transfer:  float   # max(w)+max(din)+max(dout) over cluster layers [bytes]
    l1_compute_while_transfer: float   # 2 × compute_then_transfer [bytes]


@dataclass
class Results:
    flops_per_ms: Dict[str, Dict[str, float]]         = field(default_factory=dict)
    l1_compute_then_transfer:  float                  = 0.0
    l1_compute_while_transfer: float                  = 0.0
    row_latencies:   Dict[str, Dict[int, float]]      = field(default_factory=dict)
    block_latencies: Dict[str, List[Dict]]            = field(default_factory=dict)
    total_latency:   Dict[str, float]                 = field(default_factory=dict)
    blocks:          List[Block]                      = field(default_factory=list)
    # Serialized split (seq/ctt: wall = compute + transfer).
    clusters:        Dict[str, List[ClusterSplit]]    = field(default_factory=dict)
    # Overlapped split (cwt: wall = max(compute, transfer)).
    clusters_overlap: Dict[str, List[ClusterSplit]]   = field(default_factory=dict)
    # Best (minimised) per-cluster wall = pipeline throughput, after rebalancing
    # the work over the fixed cluster count.  Keyed by pool, one per wall model.
    best_throughput:         Dict[str, float]         = field(default_factory=dict)  # seq/ctt
    best_throughput_overlap: Dict[str, float]         = field(default_factory=dict)  # cwt
    # Original per-cluster budget from config.json (for the "[config budget …]" note).
    throughput_budget:       float                    = 0.0


# ---------------------------------------------------------------------------
# Cell value helpers
# ---------------------------------------------------------------------------

EXCEL_ERRORS = ('#NAME?', '#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NULL!', '#NUM!', '#ERROR!')


def _is_blank(x) -> bool:
    if x is None:
        return True
    if isinstance(x, float) and math.isnan(x):
        return True
    return str(x).strip() == ''


def _is_excel_error(x) -> bool:
    return isinstance(x, str) and x.strip() in EXCEL_ERRORS


def _to_str(x) -> str:
    return '' if _is_blank(x) else str(x).strip()


def _to_float(x) -> float:
    if _is_blank(x):
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s in EXCEL_ERRORS:
        raise ValueError(f"cell holds Excel error {s!r}")
    return float(s)


def _to_int(x) -> int:
    return int(_to_float(x))


def col_letter(idx: int) -> str:
    """0-based column index -> Excel letter (A, B, ..., Z, AA, AB, ...)."""
    s = ''
    n = idx
    while True:
        s = chr(ord('A') + n % 26) + s
        n = n // 26 - 1
        if n < 0:
            return s


# ---------------------------------------------------------------------------
# Template loading
# ---------------------------------------------------------------------------

def read_grid(path: str) -> List[List]:
    """Read a template into a row-major 2D list, returning calculated values for formulas."""
    ext = Path(path).suffix.lower()

    if ext == '.csv':
        with open(path, newline='') as f:
            return list(csv.reader(f))

    if ext in ('.xlsx', '.xlsm'):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]

    if ext == '.ods':
        import pandas as pd
        df = pd.read_excel(path, engine='odf', sheet_name=0, header=None)
        df = df.where(df.notnull(), None)
        return df.values.tolist()

    raise ValueError(f"Unsupported file extension: {ext}. Use .csv, .xlsx, .xlsm or .ods")


def find_header_row(grid: List[List]) -> int:
    """Return the 0-based index of the header row.

    The header row is the first row in which *every* named field of
    HEADER_ALIASES has a matching label cell (same test build_column_map uses).
    Secondary sub-header rows (e.g. "Data_in (Bytes)") do not match because the
    strings differ, so they are skipped automatically.
    """
    for r, row in enumerate(grid):
        if all(
            any(_to_str(v) in aliases for v in row)
            for aliases in HEADER_ALIASES.values()
        ):
            return r
    raise ValueError(
        "Header row not found: no row labels all of "
        "N_repetitions / FLOPs / Weights / Data_in / Data_out. "
        f"(accepted aliases: {HEADER_ALIASES})"
    )


def find_data_end(grid: List[List], cols: Dict[str, int], first_row: int) -> int:
    """Return the 0-based row index where the layer table ends (exclusive).

    Layers run from `first_row` until the trailing parameter block: reading
    stops at the first row whose block-name cell (column A) is filled but whose
    operation cell (column C) is empty.  A real block header always carries its
    first layer's operation on the same row, so this only fires on the bottom
    parameter table (or a stray END-like row).  Falls back to end-of-grid.
    """
    bcol, ocol = cols['block'], cols['op']
    for r in range(first_row, len(grid)):
        row = grid[r]
        block = _to_str(row[bcol]) if bcol < len(row) else ''
        op    = _to_str(row[ocol]) if ocol < len(row) else ''
        if block and not op:
            return r
    return len(grid)


def build_column_map(header_row: List) -> Dict[str, int]:
    """Build {field -> column index}, header-based for named fields, defaults for the rest."""
    cols = dict(DEFAULT_COLS)
    for field_name, aliases in HEADER_ALIASES.items():
        match = next(
            (c for c, v in enumerate(header_row) if _to_str(v) in aliases),
            None,
        )
        if match is None:
            raise ValueError(
                f"Header for {field_name!r} not found (expected one of {aliases}). "
                f"Got header row: {[_to_str(x) for x in header_row]}"
            )
        cols[field_name] = match
    return cols


def check_for_errors(grid: List[List], cols: Dict[str, int],
                     first_row: int, last_row: int) -> None:
    """Raise a clear ValueError listing any Excel error cells in the data columns."""
    bad: List[str] = []
    for r in range(first_row, last_row + 1):
        if r >= len(grid):
            break
        row = grid[r]
        for field_name, c in cols.items():
            if c < len(row) and _is_excel_error(row[c]):
                bad.append(f"  {col_letter(c)}{r+1}  ({field_name}) = {row[c]!r}")
    if bad:
        raise ValueError(
            "Template contains unrecalculated formula errors in the data region:\n"
            + "\n".join(bad)
            + "\n\nOpen the file in Excel/LibreOffice, recalculate (Ctrl+Shift+F9), save, and try again."
        )


def parse_layers(grid: List[List], cols: Dict[str, int],
                 first_row: int, last_row: int) -> List[LayerRow]:
    """Parse data rows in the inclusive 0-based range [first_row, last_row]."""
    max_col = max(cols.values())
    layers: List[LayerRow] = []
    for r in range(first_row, last_row + 1):
        if r >= len(grid):
            break
        row = list(grid[r])
        if len(row) <= max_col:
            row += [None] * (max_col + 1 - len(row))

        op = _to_str(row[cols['op']])
        if not op or _is_blank(row[cols['flops']]):
            continue  # skip blank/spacer rows inside the region

        layers.append(LayerRow(
            excel_row     = r + 1,  # back to 1-indexed for messages / printing
            block         = _to_str(row[cols['block']]),
            n_rep         = _to_int(row[cols['n_rep']]),
            op            = op,
            flops         = _to_float(row[cols['flops']]),
            weights_bytes = _to_float(row[cols['weights']]) * BYTES_PER_ELEM,
            din_bytes     = _to_float(row[cols['din']])     * BYTES_PER_ELEM,
            dout_bytes    = _to_float(row[cols['dout']])    * BYTES_PER_ELEM,
            is_gemm       = _to_str(row[cols['gemm']]).upper() == 'GEMM',
        ))
    return layers


def group_into_blocks(rows: List[LayerRow]) -> List[Block]:
    """Group consecutive rows under their preceding block-header row."""
    blocks: List[Block] = []
    current: Optional[Block] = None
    for r in rows:
        if r.block:
            current = Block(name=r.block, n_rep=r.n_rep)
            blocks.append(current)
        if current is None:
            raise ValueError(f"Row {r.excel_row} has no preceding block header")
        current.rows.append(r)
    return blocks


# ---------------------------------------------------------------------------
# Computation
# ---------------------------------------------------------------------------

def _transfer_ms(num_bytes: float, bandwidth_gbs: Optional[float]) -> float:
    """Transfer time [ms] to move num_bytes at bandwidth_gbs GB/s.

    bandwidth_gbs GB/s == bandwidth_gbs*1e9 B/s, so ms = bytes/(bw*1e9)*1e3
    = bytes/(bw*1e6).  Returns 0.0 when bandwidth is unset/zero (compute-only).
    """
    if not bandwidth_gbs:
        return 0.0
    return num_bytes / (bandwidth_gbs * 1e6)


def split_into_clusters(
    blocks: List[Block],
    row_latencies: Dict[int, float],
    throughput_ms: float,
    bandwidth_gbs: Optional[float] = None,
    overlap: bool = False,
) -> List[List[ClusterSegment]]:
    """Greedy left-to-right partition at single-repetition granularity.

    Iterates rep-by-rep through each block (one full pass through the block's
    rows is the atomic unit).  A new cluster is opened when one more pass
    would push the cluster's wall-clock past throughput_ms.  A single oversized
    pass is always emitted as a solo entry (cannot split mid-pass).

    The per-cluster wall-clock budget depends on ``overlap``:

    * ``overlap=False`` (sequential / compute-then-transfer): compute and DMA
      are serialized, so wall = compute + transfer.
    * ``overlap=True`` (compute-while-transfer, double-buffered): compute and
      DMA overlap, so the steady-state wall = max(compute, transfer) -- a
      less conservative budget that generally yields fewer, larger clusters.

    Transfer is the hand-off cost: the prospective last block's data_out / BW,
    plus (for cluster 0 only) the HBM data_in load of the very first layer.
    With bandwidth_gbs unset the transfer term is 0, so both budgets reduce to
    the same compute-only behaviour.
    """
    din0_ms = _transfer_ms(blocks[0].rows[0].din_bytes, bandwidth_gbs) if blocks else 0.0
    clusters: List[List[ClusterSegment]] = []
    current:  List[ClusterSegment] = []
    running = 0.0
    is_first_cluster = True
    for b in blocks:
        block_sum  = sum(row_latencies[r.excel_row] for r in b.rows)
        # Hand-off to the next cluster = this block's LAST layer's data_out (e.g. the
        # skip-connection / final output). Intermediate per-row douts (Q.K^T, softmax,
        # ...) are tensors that live in the cluster's L1 and never cross the NoC, so
        # they must NOT be counted. This matches the generator's CLUSTER_DOUT_BYTES.
        xfer_ms    = _transfer_ms(b.rows[-1].dout_bytes, bandwidth_gbs)
        rep_start: Optional[int] = None
        for rep in range(b.n_rep):
            # prospective wall-clock if this pass joins the current cluster and
            # the cluster ends here: compute so far + this pass, against its
            # hand-off transfer (overlap=max, otherwise serialized=sum).
            compute_term  = running + block_sum
            transfer_term = xfer_ms + (din0_ms if is_first_cluster else 0.0)
            prospective   = (max(compute_term, transfer_term) if overlap
                             else compute_term + transfer_term)
            if running > 0 and prospective > throughput_ms:
                # close partial segment for this block before cutting
                if rep_start is not None:
                    current.append(ClusterSegment(block=b, rep_from=rep_start, rep_to=rep - 1))
                    rep_start = None
                clusters.append(current)
                current = []
                running = 0.0
                is_first_cluster = False
            if rep_start is None:
                rep_start = rep
            running += block_sum
        if rep_start is not None:
            current.append(ClusterSegment(block=b, rep_from=rep_start, rep_to=b.n_rep - 1))
    if current:
        clusters.append(current)
    return clusters


def _build_cluster_splits(
    cluster_groups: List[List[ClusterSegment]],
    row_lat: Dict[int, float],
    bandwidth_gbs: Optional[float],
) -> List[ClusterSplit]:
    """Turn a greedy cluster grouping into ClusterSplit metadata records.

    Shared by both the serialized and overlapped splits: derives per-cluster
    compute latency, hand-off transfer time, first/last layer byte sizes and
    the L1 footprints.  The grouping itself decides where the cuts fall; this
    only annotates it.
    """
    splits: List[ClusterSplit] = []
    for idx, segs in enumerate(cluster_groups):
        all_rows    = [r for seg in segs for r in seg.block.rows]
        grp_latency = sum(
            seg.n_reps * sum(row_lat[r.excel_row] for r in seg.block.rows)
            for seg in segs
        )
        max_w    = max(r.weights_bytes for r in all_rows)
        max_din  = max(r.din_bytes     for r in all_rows)
        max_dout = max(r.dout_bytes    for r in all_rows)
        l1_ctt   = max_w + max_din + max_dout
        first_r  = segs[0].block.rows[0]
        last_r   = segs[-1].block.rows[-1]
        # hand-off bytes: the cluster's pipeline output is its last layer's
        # data_out (== cs.last_dout, == the generator's CLUSTER_DOUT_BYTES).
        # Plus, on cluster 0 only, the HBM data_in load of the first layer.
        xfer_ms  = _transfer_ms(last_r.dout_bytes, bandwidth_gbs)
        if idx == 0:
            xfer_ms += _transfer_ms(first_r.din_bytes, bandwidth_gbs)
        splits.append(ClusterSplit(
            idx                       = idx,
            segments                  = segs,
            latency                   = grp_latency,
            transfer_ms               = xfer_ms,
            first_weights             = first_r.weights_bytes,
            first_din                 = first_r.din_bytes,
            last_weights              = last_r.weights_bytes,
            last_dout                 = last_r.dout_bytes,
            l1_compute_then_transfer  = l1_ctt,
            l1_compute_while_transfer = 2 * l1_ctt,
        ))
    return splits


def _makespan(splits: List[ClusterSplit], overlap: bool) -> float:
    """Pipeline bottleneck = the largest per-cluster wall over the split.

    The steady-state throughput of the pipeline is set by its slowest stage, so
    this max IS the throughput we want to minimise.  Uses the same wall model as
    ``_print_cluster_table``: max(compute, transfer) when overlapped (cwt), else
    compute + transfer (seq/ctt).
    """
    if not splits:
        return 0.0
    return max(
        (max(cs.latency, cs.transfer_ms) if overlap else cs.latency + cs.transfer_ms)
        for cs in splits
    )


def optimize_throughput(
    blocks: List[Block],
    row_lat: Dict[int, float],
    n_target: int,
    bandwidth_gbs: Optional[float],
    overlap: bool,
    budget_ms: float,
    iters: int = 60,
) -> Tuple[float, List[ClusterSplit]]:
    """Lower the per-cluster wall budget as far as possible while keeping the
    cluster count fixed at ``n_target``, then return the rebalanced split.

    Reuses ``split_into_clusters`` as a feasibility oracle: ``K(T)`` = the number
    of clusters the greedy needs for budget ``T`` is non-increasing in ``T``.
    ``n_target`` is fixed at the original ``budget_ms`` split, so for every
    ``T <= budget_ms`` we have ``K(T) >= n_target``; the smallest ``T`` with
    ``K(T) <= n_target`` therefore has ``K(T) == n_target`` exactly -- the cluster
    count is preserved automatically.  Binary-searching that smallest feasible
    ``T`` minimises the bottleneck (= evenly spreads the work).

    Returns ``(best_throughput, balanced_splits)`` where ``best_throughput`` is the
    *actual* makespan of the returned split (so a single unsplittable, dominating
    pass is reported honestly even if the search budget dipped below it).
    """
    def feasible(thr: float) -> bool:
        return len(split_into_clusters(blocks, row_lat, thr,
                                       bandwidth_gbs, overlap)) <= n_target

    # hi is feasible by construction (K(budget_ms) == n_target); lo is the
    # infeasible side.  ~60 halvings drive the gap well below FP noise; each
    # probe is O(total passes), so the whole search is cheap.
    lo, hi = 0.0, budget_ms
    for _ in range(iters):
        mid = 0.5 * (lo + hi)
        if feasible(mid):
            hi = mid
        else:
            lo = mid

    balanced = _build_cluster_splits(
        split_into_clusters(blocks, row_lat, hi, bandwidth_gbs, overlap),
        row_lat, bandwidth_gbs)

    # Safety net: never hand back a different cluster count than requested. If a
    # numerical edge (or any non-monotonic boundary) slipped through, fall back
    # to the original budget split so N is guaranteed constant.
    if len(balanced) != n_target:
        balanced = _build_cluster_splits(
            split_into_clusters(blocks, row_lat, budget_ms, bandwidth_gbs, overlap),
            row_lat, bandwidth_gbs)

    return _makespan(balanced, overlap), balanced


def compute(layers: List[LayerRow], cfg: dict) -> Results:
    util  = cfg['PE_utilization']
    pools = list(cfg['PE_peak'].keys())
    res   = Results()

    # K1:M3 -- FLOPs/ms
    for p in pools:
        res.flops_per_ms[p] = {
            'non_GEMM': cfg['PE_peak'][p]   * util * 1e9,
            'GEMM':     cfg['GEMM_peak'][p]        * 1e9,
        }

    # Y5:Z53 -- per-row, per-block and total latency [ms]
    for p in pools:
        rate_g  = res.flops_per_ms[p]['GEMM']
        rate_ng = res.flops_per_ms[p]['non_GEMM']
        res.row_latencies[p] = {
            r.excel_row: r.flops / (rate_g if r.is_gemm else rate_ng)
            for r in layers
        }

    blocks = group_into_blocks(layers)
    res.blocks = blocks
    for p in pools:
        res.block_latencies[p] = []
        total = 0.0
        for b in blocks:
            block_sum = sum(res.row_latencies[p][r.excel_row] for r in b.rows)
            contrib   = block_sum * b.n_rep
            res.block_latencies[p].append({
                'name':     b.name,
                'n_rep':    b.n_rep,
                'sum_rows': block_sum,
                'subtotal': contrib,
            })
            total += contrib
        res.total_latency[p] = total

    # Pipeline cluster splitting (skipped when throughput absent from config).
    # Build two splits per pool: the serialized budget (compute + transfer) used
    # by seq/ctt, and the overlapped budget (max(compute, transfer)) used by cwt,
    # which generally packs more work per cluster.
    if 'throughput' in cfg:
        throughput_ms = float(cfg['throughput'])
        bandwidth_gbs = cfg.get('bandwidth_gbs')
        res.throughput_budget = throughput_ms
        for p in pools:
            row_lat = res.row_latencies[p]
            # Two wall models: serialized (seq/ctt) and overlapped (cwt).  For
            # each, the existing greedy "split levels over clusters" fixes the
            # cluster count N at the budget; optimize_throughput then rebalances
            # the work within those N clusters to push the bottleneck (= the
            # pipeline throughput) as low as it will go.
            for overlap, clusters_attr, best_attr in (
                (False, res.clusters,         res.best_throughput),
                (True,  res.clusters_overlap, res.best_throughput_overlap),
            ):
                n_target = len(split_into_clusters(
                    blocks, row_lat, throughput_ms, bandwidth_gbs, overlap))
                best_tp, balanced = optimize_throughput(
                    blocks, row_lat, n_target, bandwidth_gbs, overlap, throughput_ms)
                clusters_attr[p] = balanced
                best_attr[p]     = best_tp

    # T52:W54 -- L1 footprint (bytes)
    max_w    = max(r.weights_bytes for r in layers)
    max_din  = max(r.din_bytes     for r in layers)
    max_dout = max(r.dout_bytes    for r in layers)
    res.l1_compute_then_transfer  = max_w + max_din + max_dout
    res.l1_compute_while_transfer = 2 * res.l1_compute_then_transfer

    return res


# ---------------------------------------------------------------------------
# Printing
# ---------------------------------------------------------------------------

def _print_cluster_table(pool: str, clusters: List[ClusterSplit],
                         row_lat: Dict[int, float], *, overlap: bool,
                         best_throughput: float, budget: float) -> None:
    """Print the per-cluster split table for one pool under one wall model.

    overlap selects the wall-clock budget: max(compute, transfer) for the
    compute-while-transfer (cwt) schedule, else compute + transfer (seq/ctt).
    best_throughput is the minimised bottleneck after rebalancing the work over
    the fixed cluster count; it is printed prominently at the head of the table.
    """
    label = ('Overlapped  (cwt: compute-while-transfer)' if overlap
             else 'Serialized  (seq / ctt: compute-then-transfer)')
    model = 'max(compute, transfer)' if overlap else 'compute + transfer'
    n_clusters = len(clusters)
    print(f'\nPool: {pool}  --  {label}')
    print(f'  >>> Best throughput found = {best_throughput:.6g} ms  '
          f'(min bottleneck over {n_clusters} clusters held constant; '
          f'config budget {budget:.6g} ms)')
    print(f'  {n_clusters} cluster(s)  |  wall = {model}')
    for cs in clusters:
        print('  ' + '=' * 78)
        seg_desc = ', '.join(
            f'{seg.block.name} rep {seg.rep_from+1}-{seg.rep_to+1}/{seg.block.n_rep}'
            for seg in cs.segments
        )
        wall = max(cs.latency, cs.transfer_ms) if overlap else cs.latency + cs.transfer_ms
        print(f'  Cluster {cs.idx}  |  compute = {cs.latency:.6g} ms  |  '
              f'transfer = {cs.transfer_ms:.6g} ms  |  wall = {wall:.6g} ms')
        print(f'  {"":11}{seg_desc}')
        print(f'  {"":4}{"Block":<16}{"Reps":<8}{"Row":<6}{"Op":<26}{"":6}{"lat/rep [ms]":>14}{"×N":>4}{"total [ms]":>12}')
        print('  ' + '-' * 78)
        for seg in cs.segments:
            rep_tag = f'{seg.rep_from+1}-{seg.rep_to+1}' if seg.n_reps > 1 else f'{seg.rep_from+1}'
            for r in seg.block.rows:
                tag     = ' GEMM' if r.is_gemm else '     '
                lat     = row_lat[r.excel_row]
                tot_lat = lat * seg.n_reps
                print(f'  {"":4}{seg.block.name[:16]:<16}{rep_tag:<8}{r.excel_row:<6}'
                      f'{r.op[:26]:<26}{tag}  {lat:>14.6g}{seg.n_reps:>4}{tot_lat:>12.6g}')
        print('  ' + '-' * 78)
        print(f'  {"":4}First layer : weights={cs.first_weights:>12,.0f} B'
              f'  data_in ={cs.first_din:>12,.0f} B')
        if cs.idx < n_clusters - 1:
            print(f'  {"":4}Last  layer : weights={cs.last_weights:>12,.0f} B'
                  f'  data_out={cs.last_dout:>11,.0f} B  --> pipeline to cluster {cs.idx+1}')
        else:
            print(f'  {"":4}Last  layer : weights={cs.last_weights:>12,.0f} B'
                  f'  data_out={cs.last_dout:>11,.0f} B')
        print(f'  {"":4}Max L1 compute-then-transfer  = {cs.l1_compute_then_transfer:>12,.0f} B')
        print(f'  {"":4}Max L1 compute-while-transfer = {cs.l1_compute_while_transfer:>12,.0f} B')
    print('  ' + '=' * 74)


def print_results(layers: List[LayerRow], res: Results,
                  mode: Optional[str] = None) -> None:
    pools = list(res.flops_per_ms.keys())

    print('#' * 78)
    print('FLOPs/ms  (achievable FLOPs per millisecond)')
    print('#' * 78)
    print(f'{"":<14}{"non-GEMM":>20}{"GEMM":>20}')
    for pool, vals in res.flops_per_ms.items():
        print(f'{pool:<14}{vals["non_GEMM"]:>20,.0f}{vals["GEMM"]:>20,.0f}')

    print()
    print('#' * 78)
    print('L1 footprint [bytes]')
    print('#' * 78)
    print(f'Compute then transfer :  {res.l1_compute_then_transfer:>15,.0f}')
    print(f'Compute while transfer:  {res.l1_compute_while_transfer:>15,.0f}')

    print()
    print('#' * 78)
    print('Latency [ms] per pool')
    print('#' * 78)
    header = f'{"Row":<5}{"Block":<16}{"Op":<20}{"":<5}' + ''.join(f'{p:>16}' for p in pools)
    print(header)
    print('-' * len(header))

    blocks = res.blocks
    for b in blocks:
        for r in b.rows:
            tag  = 'GEMM' if r.is_gemm else '   .'
            vals = ''.join(f'{res.row_latencies[p][r.excel_row]:>16.6g}' for p in pools)
            print(f'{r.excel_row:<5}{b.name:<16}{r.op[:20]:<20}{tag:<5}{vals}')
        subtotal = ''.join(
            f'{next(x for x in res.block_latencies[p] if x["name"] == b.name)["subtotal"]:>16.6g}'
            for p in pools
        )
        print(f'{"":<5}{"":<16}{f"-> x{b.n_rep} subtotal":<25}{subtotal}')
        print()

    print('-' * len(header))
    totals = ''.join(f'{res.total_latency[p]:>16.6g}' for p in pools)
    print(f'{"":<5}{"":<16}{"Tot min latency":<25}{totals}')

    if res.clusters:
        print()
        print('#' * 78)
        print('Pipeline cluster split')
        print('#' * 78)
        # Show only the split relevant to the chosen --mode: serialized
        # (compute + transfer) for seq/ctt, overlapped (max(compute, transfer))
        # for cwt.  With no --mode show both for inspection.
        show_serial  = mode in (None, 'seq', 'ctt')
        show_overlap = mode in (None, 'cwt')
        for p in pools:
            if show_serial and p in res.clusters:
                _print_cluster_table(p, res.clusters[p],
                                     res.row_latencies[p], overlap=False,
                                     best_throughput=res.best_throughput[p],
                                     budget=res.throughput_budget)
            if show_overlap and p in res.clusters_overlap:
                _print_cluster_table(p, res.clusters_overlap[p],
                                     res.row_latencies[p], overlap=True,
                                     best_throughput=res.best_throughput_overlap[p],
                                     budget=res.throughput_budget)


# ---------------------------------------------------------------------------
# Shared test-generation renderers
# ---------------------------------------------------------------------------

def _render_cmake_txt() -> str:
    return (
        "set(SRC_SOURCES ${CMAKE_CURRENT_SOURCE_DIR}/main.c)\n"
        "set(SOURCES ${SRC_SOURCES} PARENT_SCOPE)\n"
        "set(INCLUDE_DIRS ${CMAKE_CURRENT_SOURCE_DIR}/include PARENT_SCOPE)\n"
    )


def _render_dma_helper() -> str:
    """Shared C: AXI-pool notes, DMA size macros, and the blocking chunked-DMA helper.

    Reused verbatim by the sequential generator and the pipeline generators.
    The 7-page / DMA_MAX_BYTES cap must not change: it is dictated by the GVSoC
    iDMA AXI burst-queue pool.
    """
    return '\n'.join([
        '/*',
        ' * The GVSoC iDMA AXI backend has a fixed pool of burst_queue_size(8) buffers of',
        ' * AXI_PAGE_SIZE(4096) B each, shared by reads AND writes. A single transfer is',
        ' * page-split on its AXI-side address, so it can span at most 8 pages before the',
        ' * pool is exhausted and flex_dma_async_wait_all() hangs. The usable byte count is',
        " * reduced by the AXI-side address's offset within its first 4 KB page.",
        ' *',
        ' * dma_1d_chunked() caps every chunk at 7 pages (28672 B). From ANY page offset,',
        ' * 28672 B spans <= 8 pages, so it never exhausts the pool -- safe regardless of',
        ' * alignment or src/dst direction.',
        ' */',
        '#define DMA_BUS_BYTES   64u      /* keep issued sizes a multiple of the bus width */',
        '#define DMA_MAX_BYTES   28672u   /* 7 * AXI_PAGE_SIZE(4096): <= 8 pages from any offset */',
        '#define DMA_ALIGN_UP(n) (((n) + (DMA_BUS_BYTES - 1u)) & ~(DMA_BUS_BYTES - 1u))',
        '',
        '/* Blocking 1D DMA: every issued chunk is <= DMA_MAX_BYTES and a multiple of',
        ' * DMA_BUS_BYTES. Non-final chunks are DMA_MAX_BYTES (already 64-aligned); the final',
        ' * remainder is rounded up to DMA_BUS_BYTES. The transfer loop is bracketed with',
        ' * mempool_get_timer() (mcycle CSR) so DMA time can be separated from compute. */',
        'static inline void dma_1d_chunked(uint64_t dst, uint64_t src, uint32_t bytes)',
        '{',
        '    mempool_timer_t t0 = mempool_get_timer();',
        '    while (bytes) {',
        '        uint32_t n = (bytes >= DMA_MAX_BYTES) ? DMA_MAX_BYTES : DMA_ALIGN_UP(bytes);',
        '        flex_dma_async_1d(dst, src, n);',
        '        flex_dma_async_wait_all();',
        '        if (bytes <= DMA_MAX_BYTES) break;   /* last (aligned) chunk issued */',
        '        // dst += DMA_MAX_BYTES; src += DMA_MAX_BYTES;',
        '        bytes -= DMA_MAX_BYTES;',
        '    }',
        '    uint32_t dt = mempool_get_timer() - t0;',
        '    printf("[dma] %u cycles \\n", dt);',
        '}',
    ])


# ---------------------------------------------------------------------------
# Pipeline (ctt/cwt) per-layer compute helpers
# ---------------------------------------------------------------------------

def _layer_cycles(row: 'LayerRow', row_lat: Dict[int, float],
                  cycles_per_ms: int) -> int:
    """Per-layer wait-loop count (>=1) from the row's latency [ms]."""
    return max(1, int(row_lat.get(row.excel_row, 0.0) * cycles_per_ms))


def _c_comment(text: str) -> str:
    return text.replace('*/', '* /')


def _render_compute_body(cluster: 'ClusterSplit', row_lat: Dict[int, float],
                         cycles_per_ms: int) -> List[str]:
    """One `mempool_wait` per layer, block reps wrapped in a `for` loop."""
    lines: List[str] = []
    for seg in cluster.segments:
        b   = seg.block
        rng = (f'{seg.rep_from+1}-{seg.rep_to+1}/{b.n_rep}'
               if seg.n_reps > 1 else f'{seg.rep_from+1}/{b.n_rep}')
        lines.append(f'    /* {_c_comment(b.name)} x{seg.n_reps}  (reps {rng}) */')
        lines.append(f'    for (uint32_t r = 0; r < {seg.n_reps}u; r++) {{')
        for row in b.rows:
            cyc = _layer_cycles(row, row_lat, cycles_per_ms)
            tag = ' (GEMM)' if row.is_gemm else ''
            lines.append(f'        mempool_wait({cyc}UL);  '
                         f'/* {_c_comment(b.name)}: {_c_comment(row.op)}{tag} */')
        lines.append('    }')
    return lines


# ---------------------------------------------------------------------------
# C renderers (one per mode)
# ---------------------------------------------------------------------------

def _render_seq_c(pool: str, clusters: List[ClusterSplit], cycles_per_ms: int) -> str:
    n             = len(clusters)
    first_din     = int(clusters[0].first_din)
    max_transfer  = max(first_din, max(int(cs.last_dout) for cs in clusters))
    # round buffer up to a 64-byte boundary (32 FP16 elements per 64 bytes)
    max_elems     = ((max_transfer + 63) // 64) * 32
    hbm_out_off   = (first_din + 63) & ~63   # HBM output slot, aligned after input

    lines: List[str] = []
    A = lines.append

    A(f'/* Auto-generated by pipeline.py -- Pipeline simulation test for {pool} */')
    A(f'/* {n} pipeline stage(s), one stage per hardware cluster. */')
    A('')
    A('#include "flex_runtime.h"')
    A('#include "flex_dma_pattern.h"')
    A('#include "flex_printf.h"')
    A('')
    A('/*')
    A(' * CYCLES_PER_MS: nop-loop iterations that approximate 1 ms of wall time.')
    A(' * Tune to your simulation / target clock frequency.')
    A(' */')
    A(f'#define CYCLES_PER_MS  {cycles_per_ms}UL')
    A('')
    A(f'#define N_PIPELINE_CLUSTERS  {n}')
    A(f'#define CLUSTER_0_DIN_BYTES  {first_din}U  /* first layer data_in  */')
    A(f'#define HBM_INPUT_OFFSET     0UL')
    A(f'#define HBM_OUTPUT_OFFSET    {hbm_out_off}UL  /* after input data in HBM */')
    A('')
    A(_render_dma_helper())
    A('')
    A('/* Compute wait-loop count per cluster (latency_ms x CYCLES_PER_MS) */')
    A(f'static const uint32_t CLUSTER_LATENCY_CYCLES[N_PIPELINE_CLUSTERS] = {{')
    for i, cs in enumerate(clusters):
        cyc = max(1, int(cs.latency * cycles_per_ms))
        A(f'    {cyc}UL,  /* cluster {i}: {cs.latency:.6g} ms */')
    A('};')
    A('')
    A('/* data_out transfer size per cluster [bytes] (= bytes passed to next stage / HBM) */')
    A(f'static const uint32_t CLUSTER_DOUT_BYTES[N_PIPELINE_CLUSTERS] = {{')
    for i, cs in enumerate(clusters):
        seg_info = ', '.join(
            f'{seg.block.name}({seg.rep_from+1}-{seg.rep_to+1}/{seg.block.n_rep})'
            for seg in cs.segments
        )
        A(f'    {int(cs.last_dout)}U,  /* cluster {i}: [{seg_info}] */')
    A('};')
    A('')
    A('/* L1 pipeline buffer -- sized for the largest single transfer */')
    A(f'#define L1_BUF_ELEMS  {max_elems}U  /* = {max_elems * 2} bytes (FP16) */')
    A(f'static uint16_t l1_buf[L1_BUF_ELEMS]')
    A(f'    __attribute__((section(".l1"), aligned(64)));')
    A('')
    A('int main()')
    A('{')
    A('    uint32_t cid     = flex_get_cluster_id();')
    A('    uint32_t eoc_val = 0;')
    if n > 1:
        A('    /* buf_off: TCDM offset of l1_buf -- identical in every cluster (same binary) */')
        A('    const uint32_t buf_off = (uint32_t)(uintptr_t)l1_buf;')
    A('')
    A('    flex_barrier_xy_init();')
    A('    flex_global_barrier_xy();')
    A('    if (flex_get_core_id() == 0 && cid == 0) flex_timer_start();')
    A('    flex_global_barrier_xy();')
    A('')
    A('    /**************************************/')
    A('    /*  Program Execution Region -- Start */')
    A('    /**************************************/')
    A('')
    A('    /*')
    A('     * Each pipeline stage runs on one hardware cluster (cid == stage index).')
    A('     * Only the DM core (flex_is_dm_core()) of each active cluster acts.')
    A('     * Stages execute sequentially, separated by global barriers.')
    A('     *')
    A('     * Data flow:')
    A('     *   HBM ---[DMA]--> Cluster 0 --[DMA]--> Cluster 1 --> ... --> Cluster N-1 --[DMA]--> HBM')
    A('     */')
    A('')

    for i, cs in enumerate(clusters):
        is_first = (i == 0)
        is_last  = (i == n - 1)
        seg_str  = ', '.join(f'{seg.block.name}x{seg.n_reps}' for seg in cs.segments)

        A(f'    /* ---- Stage {i} | [{seg_str}] | latency ~ {cs.latency:.4g} ms ---- */')
        A(f'    if (flex_is_dm_core() && cid == {i}) {{')
        if is_first:
            A(f'        /* Load first-cluster data_in from HBM */')
            A(f'        printf("[Cluster {i}] Load %u B from HBM\\n", CLUSTER_0_DIN_BYTES);')
            A(f'        dma_1d_chunked((uint64_t)(uintptr_t)l1_buf,')
            A(f'                       hbm_addr(HBM_INPUT_OFFSET),')
            A(f'                       CLUSTER_0_DIN_BYTES);')
        A(f'        /* Fake compute: spin for ~ {cs.latency:.4g} ms */')
        A(f'        mempool_wait(CLUSTER_LATENCY_CYCLES[{i}]);')
        if is_last:
            A(f'        /* Store last-cluster data_out to HBM */')
            A(f'        printf("[Cluster {i}] Store %u B to HBM\\n", CLUSTER_DOUT_BYTES[{i}]);')
            A(f'        dma_1d_chunked(hbm_addr(HBM_OUTPUT_OFFSET),')
            A(f'                       (uint64_t)(uintptr_t)l1_buf,')
            A(f'                       CLUSTER_DOUT_BYTES[{i}]);')
        else:
            A(f'        /* Push data_out to Cluster {i + 1} L1 */')
            A(f'        printf("[Cluster {i}] Push %u B to cluster {i + 1}\\n", CLUSTER_DOUT_BYTES[{i}]);')
            A(f'        dma_1d_chunked((uint64_t)remote_cid({i + 1}, buf_off),')
            A(f'                       (uint64_t)(uintptr_t)l1_buf,')
            A(f'                       CLUSTER_DOUT_BYTES[{i}]);')
        A('    }')
        A('    flex_global_barrier_xy();')
        A('')

    A('    /**************************************/')
    A('    /*  Program Execution Region -- Stop  */')
    A('    /**************************************/')
    A('    if (flex_get_core_id() == 0 && cid == 0) flex_timer_end();')
    A('    flex_global_barrier_xy();')
    A('    flex_eoc(eoc_val);')
    A('    return 0;')
    A('}')

    return '\n'.join(lines) + '\n'


def _render_ctt_c(pool: str, clusters: List[ClusterSplit],
                  row_lat: Dict[int, float], cycles_per_ms: int, n_iter: int) -> str:
    n            = len(clusters)
    first_din    = int(clusters[0].first_din)
    max_transfer = max(first_din, max(int(cs.last_dout) for cs in clusters))
    # round buffer up to a 64-byte boundary (32 FP16 elements per 64 bytes)
    max_elems    = ((max_transfer + 63) // 64) * 32
    hbm_out_off  = (first_din + 63) & ~63

    lines: List[str] = []
    A = lines.append

    A(f'/* Auto-generated by pipeline.py -- compute-then-transfer pipeline for {pool} */')
    A(f'/* {n} stage(s), one per cluster; {n_iter} item(s) streamed as a wavefront. */')
    A('')
    A('#include "flex_runtime.h"')
    A('#include "flex_dma_pattern.h"')
    A('#include "flex_printf.h"')
    A('')
    A('/* CYCLES_PER_MS: nop-loop iterations that approximate 1 ms of wall time. */')
    A(f'#define CYCLES_PER_MS  {cycles_per_ms}UL')
    A('')
    A(f'#define N_PIPELINE_CLUSTERS  {n}')
    A(f'#define N_ITER               {n_iter}U  /* items streamed through the pipeline */')
    A(f'#define CLUSTER_0_DIN_BYTES  {first_din}U  /* first layer data_in  */')
    A(f'#define HBM_INPUT_OFFSET     0UL')
    A(f'#define HBM_OUTPUT_OFFSET    {hbm_out_off}UL  /* after input data in HBM */')
    A('')
    A(_render_dma_helper())
    A('')
    A('/* data_out transfer size per cluster [bytes] (= bytes passed to next stage / HBM) */')
    A('static const uint32_t CLUSTER_DOUT_BYTES[N_PIPELINE_CLUSTERS] = {')
    for i, cs in enumerate(clusters):
        seg_info = ', '.join(
            f'{seg.block.name}({seg.rep_from+1}-{seg.rep_to+1}/{seg.block.n_rep})'
            for seg in cs.segments
        )
        A(f'    {int(cs.last_dout)}U,  /* cluster {i}: [{seg_info}] */')
    A('};')
    A('')
    A('/* ---- Per-cluster compute: one mempool_wait() per layer (future real-call site). */')
    for i, cs in enumerate(clusters):
        seg_str = ', '.join(f'{seg.block.name} x{seg.n_reps}' for seg in cs.segments)
        A(f'/* stage {i} : [{_c_comment(seg_str)}] */')
        A(f'static inline void compute_stage_{i}(void)')
        A('{')
        lines.extend(_render_compute_body(cs, row_lat, cycles_per_ms))
        A('}')
        A('')
    A('/* Single L1 buffer per cluster -- compute-then-transfer (no double buffering). */')
    A(f'#define L1_BUF_ELEMS  {max_elems}U  /* = {max_elems * 2} bytes (FP16) */')
    A('static uint16_t l1_buf[L1_BUF_ELEMS]')
    A('    __attribute__((section(".l1"), aligned(64)));')
    A('')
    A('int main()')
    A('{')
    A('    uint32_t cid     = flex_get_cluster_id();')
    A('    uint32_t eoc_val = 0;')
    A('    /* buf_off: TCDM offset of l1_buf -- identical in every cluster (same binary) */')
    A('    const uint32_t buf_off = (uint32_t)(uintptr_t)l1_buf;')
    A('')
    A('    flex_barrier_xy_init();')
    A('    flex_global_barrier_xy();')
    A('    if (flex_get_core_id() == 0 && cid == 0) flex_timer_start();')
    A('    flex_global_barrier_xy();')
    A('')
    A('    /**************************************/')
    A('    /*  Program Execution Region -- Start */')
    A('    /**************************************/')
    A('')
    A('    /*')
    A('     * Real streaming pipeline (compute-then-transfer).')
    A('     * step counts pipeline ticks: cluster c handles item (step - c) when')
    A('     * 0 <= step - c < N_ITER.  The loop covers FILL (clusters switch on as step')
    A('     * grows), STEADY STATE (all N busy on different items) and DRAIN (clusters')
    A('     * switch off as items leave).  Data is placeholder -- only cycles are measured.')
    A('     *')
    A('     * Within each active cluster the worker cores run compute_stage_c() while the')
    A('     * dm_core handles DMA.  compute -> barrier -> transfer -> barrier, so compute and')
    A('     * the outbound transfer do NOT overlap (compute-then-transfer).')
    A('     *')
    A('     *   HBM --[DMA]--> C0 --[DMA]--> C1 --> ... --> C(N-1) --[DMA]--> HBM')
    A('     */')
    A('    for (uint32_t step = 0; step < N_ITER + N_PIPELINE_CLUSTERS - 1; step++) {')
    A('        uint32_t active = (cid < N_PIPELINE_CLUSTERS) && (cid <= step)')
    A('                          && (step - cid < N_ITER);')
    A('')
    A('        /* ---- Compute phase: ALL cores compute; dm_core also loads cid0 input ---- */')
    A('        if (active) {')
    A('            if (cid == 0) {')
    A('                /* Cluster 0 input arrives from HBM this step: its worker cores must')
    A('                 * not compute until the dm_core load has completed. */')
    A('                if (flex_is_dm_core()) {')
    A('                    printf("[step %u][C0] load item %u from HBM\\n", step, step);')
    A('                    dma_1d_chunked((uint64_t)(uintptr_t)l1_buf,')
    A('                                   hbm_addr(HBM_INPUT_OFFSET),')
    A('                                   CLUSTER_0_DIN_BYTES);')
    A('                }')
    A('                flex_intra_cluster_sync();   /* worker cores wait for the HBM load */')
    A('            }')
    A('            /* Every core computes (incl. dm_core): on cid0 it has finished the HBM')
    A('             * load + intra-cluster sync above, so it is free to join the compute. */')
    A('            switch (cid) {')
    for i in range(n):
        A(f'                case {i}: compute_stage_{i}(); break;')
    A('                default: break;')
    A('            }')
    A('        }')
    A('        flex_global_barrier_xy();   /* compute done before ANY transfer */')
    A('')
    A('        /* ---- Transfer phase (all active clusters in parallel) ---- */')
    A('        if (flex_is_dm_core() && active) {')
    A('            if (cid == N_PIPELINE_CLUSTERS - 1) {')
    A('                printf("[step %u][C%u] store item %u to HBM\\n", step, cid, step - cid);')
    A('                dma_1d_chunked(hbm_addr(HBM_OUTPUT_OFFSET),')
    A('                               (uint64_t)(uintptr_t)l1_buf,')
    A('                               CLUSTER_DOUT_BYTES[cid]);')
    A('            } else {')
    A('                printf("[step %u][C%u] push item %u to C%u\\n", step, cid, step - cid, cid + 1);')
    A('                dma_1d_chunked((uint64_t)remote_cid(cid + 1, buf_off),')
    A('                               (uint64_t)(uintptr_t)l1_buf,')
    A('                               CLUSTER_DOUT_BYTES[cid]);')
    A('            }')
    A('        }')
    A('        flex_global_barrier_xy();   /* all transfers done before next compute */')
    A('    }')
    A('')
    A('    /**************************************/')
    A('    /*  Program Execution Region -- Stop  */')
    A('    /**************************************/')
    A('    if (flex_get_core_id() == 0 && cid == 0) flex_timer_end();')
    A('    flex_global_barrier_xy();')
    A('    flex_eoc(eoc_val);')
    A('    return 0;')
    A('}')

    return '\n'.join(lines) + '\n'


def _render_cwt_c(pool: str, clusters: List[ClusterSplit],
                  row_lat: Dict[int, float], cycles_per_ms: int, n_iter: int) -> str:
    n            = len(clusters)
    first_din    = int(clusters[0].first_din)
    max_transfer = max(first_din, max(int(cs.last_dout) for cs in clusters))
    # round buffer up to a 64-byte boundary (32 FP16 elements per 64 bytes)
    max_elems    = ((max_transfer + 63) // 64) * 32
    hbm_out_off  = (first_din + 63) & ~63

    lines: List[str] = []
    A = lines.append

    A(f'/* Auto-generated by pipeline.py -- compute-while-transfer pipeline for {pool} */')
    A(f'/* {n} stage(s), one per cluster; {n_iter} item(s) streamed as a wavefront. */')
    A('')
    A('#include "flex_runtime.h"')
    A('#include "flex_dma_pattern.h"')
    A('#include "flex_printf.h"')
    A('')
    A('/* CYCLES_PER_MS: nop-loop iterations that approximate 1 ms of wall time. */')
    A(f'#define CYCLES_PER_MS  {cycles_per_ms}UL')
    A('')
    A(f'#define N_PIPELINE_CLUSTERS  {n}')
    A(f'#define N_ITER               {n_iter}U  /* items streamed through the pipeline */')
    A(f'#define CLUSTER_0_DIN_BYTES  {first_din}U  /* first layer data_in  */')
    A(f'#define HBM_INPUT_OFFSET     0UL')
    A(f'#define HBM_OUTPUT_OFFSET    {hbm_out_off}UL  /* after input data in HBM */')
    A('')
    A(_render_dma_helper())
    A('')
    A('/* data_out transfer size per cluster [bytes] (= bytes passed to next stage / HBM) */')
    A('static const uint32_t CLUSTER_DOUT_BYTES[N_PIPELINE_CLUSTERS] = {')
    for i, cs in enumerate(clusters):
        seg_info = ', '.join(
            f'{seg.block.name}({seg.rep_from+1}-{seg.rep_to+1}/{seg.block.n_rep})'
            for seg in cs.segments
        )
        A(f'    {int(cs.last_dout)}U,  /* cluster {i}: [{seg_info}] */')
    A('};')
    A('')
    A('/* ---- Per-cluster compute: one mempool_wait() per layer (future real-call site).')
    A('   Run by the WORKER cores while the dm_core drives the DMA in parallel. */')
    for i, cs in enumerate(clusters):
        seg_str = ', '.join(f'{seg.block.name} x{seg.n_reps}' for seg in cs.segments)
        A(f'/* stage {i} : [{_c_comment(seg_str)}] */')
        A(f'static inline void compute_stage_{i}(void)')
        A('{')
        lines.extend(_render_compute_body(cs, row_lat, cycles_per_ms))
        A('}')
        A('')
    A('/* Double-buffered L1 per cluster -- compute-while-transfer (2x footprint). */')
    A(f'#define L1_BUF_ELEMS  {max_elems}U  /* = {max_elems * 2} bytes (FP16) per buffer */')
    A('static uint16_t l1_buf[2][L1_BUF_ELEMS]')
    A('    __attribute__((section(".l1"), aligned(64)));')
    A('')
    A('int main()')
    A('{')
    A('    uint32_t cid     = flex_get_cluster_id();')
    A('    uint32_t eoc_val = 0;')
    A('')
    A('    flex_barrier_xy_init();')
    A('    flex_global_barrier_xy();')
    A('    if (flex_get_core_id() == 0 && cid == 0) flex_timer_start();')
    A('    flex_global_barrier_xy();')
    A('')
    A('    /**************************************/')
    A('    /*  Program Execution Region -- Start */')
    A('    /**************************************/')
    A('')
    A('    /*')
    A('     * Real streaming pipeline (compute-while-transfer), hazard-free double buffering.')
    A('     *')
    A('     * Two constraints set the schedule:')
    A('     *   (1) intra-cluster overlap: a cluster computes item m into one buffer while')
    A('     *       its dm_core pushes item (m-1) from the other => push lags own compute by 1.')
    A('     *   (2) inter-cluster: a consumer may compute item m only the step AFTER its')
    A('     *       producer pushed item m (the per-step barrier publishes the push).')
    A('     * Chaining (1)+(2): compute(c+1,m) = compute(c,m) + 2  =>  inter-stage offset 2.')
    A('     *')
    A('     * Buffers are indexed by ITEM PARITY.  Cluster c computes item m = step - 2*cid;')
    A('     * since 2*cid is even, cbuf = step%2 and the outbound buffer obuf = cbuf^1.  The')
    A('     * producer and consumer of item m both resolve to m%2, so the producer writes')
    A('     * exactly the buffer the consumer reads next -- no free-running index, no hazard.')
    A('     *')
    A('     * Core-role split: dm_core does ONLY the DMA (cid==0 HBM load + outbound), worker')
    A('     * cores run compute_stage_c().  One flex_global_barrier_xy() per step joins them')
    A('     * => step ~ max(compute, transfer).  Data is placeholder -- only cycles measured.')
    A('     *')
    A('     *   HBM --[DMA]--> C0 --[DMA]--> C1 --> ... --> C(N-1) --[DMA]--> HBM')
    A('     */')
    A('    for (uint32_t step = 0; step < 2u * (N_PIPELINE_CLUSTERS - 1) + N_ITER + 1u; step++) {')
    A('        uint32_t base = 2u * cid;')
    A('        /* compute item (step-2*cid) into l1_buf[cbuf]; push item (step-2*cid-1) from l1_buf[obuf] */')
    A('        uint32_t compute_active  = (cid < N_PIPELINE_CLUSTERS) && (step >= base)')
    A('                                   && (step - base < N_ITER);')
    A('        uint32_t outbound_active = (cid < N_PIPELINE_CLUSTERS) && (step >= base + 1u)')
    A('                                   && (step - base - 1u < N_ITER);')
    A('        uint32_t cbuf = step & 1u;        /* compute buffer  = item parity */')
    A('        uint32_t obuf = cbuf ^ 1u;        /* outbound buffer = previous item parity */')
    A('')
    A('        if (compute_active || outbound_active) {')
    A('            if (cid == 0 && compute_active) {')
    A('                /* Cluster 0 input arrives from HBM this step: its worker cores must')
    A('                 * not compute until the dm_core load has completed.  (The outbound')
    A('                 * transfer below still overlaps the compute.) */')
    A('                if (flex_is_dm_core()) {')
    A('                    printf("[step %u][C0] load item %u from HBM\\n", step, step - base);')
    A('                    dma_1d_chunked((uint64_t)(uintptr_t)l1_buf[cbuf],')
    A('                                   hbm_addr(HBM_INPUT_OFFSET),')
    A('                                   CLUSTER_0_DIN_BYTES);')
    A('                }')
    A('                flex_intra_cluster_sync();   /* worker cores wait for the HBM load */')
    A('            }')
    A('            if (flex_is_dm_core()) {')
    A('                /* dm_core: outbound transfer of the PREVIOUS item -- overlaps compute. */')
    A('                if (outbound_active) {')
    A('                    uint32_t item_out = step - base - 1u;')
    A('                    if (cid == N_PIPELINE_CLUSTERS - 1) {')
    A('                        printf("[step %u][C%u] store item %u to HBM\\n", step, cid, item_out);')
    A('                        dma_1d_chunked(hbm_addr(HBM_OUTPUT_OFFSET),')
    A('                                       (uint64_t)(uintptr_t)l1_buf[obuf],')
    A('                                       CLUSTER_DOUT_BYTES[cid]);')
    A('                    } else {')
    A('                        printf("[step %u][C%u] push item %u to C%u\\n", step, cid, item_out, cid + 1);')
    A('                        dma_1d_chunked((uint64_t)remote_cid(cid + 1, (uint32_t)(uintptr_t)l1_buf[obuf]),')
    A('                                       (uint64_t)(uintptr_t)l1_buf[obuf],')
    A('                                       CLUSTER_DOUT_BYTES[cid]);')
    A('                    }')
    A('                }')
    A('            } else if (compute_active) {')
    A('                /* worker cores: compute this stage\'s layers into l1_buf[cbuf]. */')
    A('                switch (cid) {')
    for i in range(n):
        A(f'                    case {i}: compute_stage_{i}(); break;')
    A('                    default: break;')
    A('                }')
    A('            }')
    A('        }')
    A('        flex_global_barrier_xy();   /* joins compute (workers) & DMA (dm_core), then clusters */')
    A('    }')
    A('')
    A('    /**************************************/')
    A('    /*  Program Execution Region -- Stop  */')
    A('    /**************************************/')
    A('    if (flex_get_core_id() == 0 && cid == 0) flex_timer_end();')
    A('    flex_global_barrier_xy();')
    A('    flex_eoc(eoc_val);')
    A('    return 0;')
    A('}')

    return '\n'.join(lines) + '\n'


# Mode -> (C renderer, banner text builder)
_C_RENDERERS = {'seq': _render_seq_c, 'ctt': _render_ctt_c, 'cwt': _render_cwt_c}


# ---------------------------------------------------------------------------
# Preload .elf builder (used by --gen-preload)
# ---------------------------------------------------------------------------

def _build_preload_elf(test_dir: Path, pool_safe: str, first_din_bytes: int) -> Path:
    """Build preload_<pool>.elf for a test dir (only used with --gen-preload).

    Random FP16 input sized to the first cluster's data_in, written to HBM base
    0xc0000000.  make_preload_elf writes
    array.c/link.ld/array.o into the CWD and shells out to riscv32-unknown-elf-*,
    so we run inside test_dir and restore the CWD afterwards.
    """
    import numpy as np
    util_dir = Path(__file__).resolve().parent.parent / 'flex_cluster_utilities'
    if str(util_dir) not in sys.path:
        sys.path.insert(0, str(util_dir))
    import preload as pld

    if shutil.which('riscv32-unknown-elf-gcc') is None:
        print('  Warning: riscv32-unknown-elf-gcc not found on PATH; '
              'the preload .elf will not be built correctly.')

    n_elems  = first_din_bytes // 2
    hbm_base = 0xc0000000
    elf_name = f'preload_{pool_safe}.elf'

    rng        = np.random.default_rng()
    input_data = rng.random(n_elems).astype(np.float16)

    cwd = os.getcwd()
    try:
        os.chdir(test_dir)
        pld.make_preload_elf(elf_name, [input_data], [hbm_base])
    finally:
        os.chdir(cwd)
    return test_dir / elf_name


# ---------------------------------------------------------------------------
# Test generation
# ---------------------------------------------------------------------------

_TESTS_DIR = Path(__file__).resolve().parent.parent.parent / 'flex_cluster_sdk' / 'tests'

DEFAULT_OUT = {
    'seq': _TESTS_DIR / '07_pipeline_seq',
    'ctt': _TESTS_DIR / '07_pipeline_sim_ctt',
    'cwt': _TESTS_DIR / '07_pipeline_sim_cwt',
}


def generate_tests(res: Results, output_dir: str, cycles_per_ms: int,
                   mode: str, n_iter: int, gen_preload: bool) -> None:
    # cwt overlaps compute & DMA, so it uses the max(compute, transfer) split;
    # seq/ctt serialize them and use the compute + transfer split.
    src = res.clusters_overlap if mode == 'cwt' else res.clusters
    if not src:
        print('Warning: no cluster splits found. '
              'Add "throughput" to config.json and re-run.')
        return

    if mode == 'seq':
        print(f'\nGenerating pipeline tests -> {output_dir}/')
    elif mode == 'ctt':
        print(f'\nGenerating compute-then-transfer pipeline tests '
              f'({n_iter} iterations) -> {output_dir}/')
    else:  # cwt
        print(f'\nGenerating compute-while-transfer pipeline tests '
              f'({n_iter} iterations) -> {output_dir}/')

    for pool, clusters in src.items():
        if not clusters:
            continue
        pool_safe = pool.replace(' ', '_')
        test_dir  = Path(output_dir) / pool_safe
        test_dir.mkdir(parents=True, exist_ok=True)

        if mode == 'seq':
            c_src = _render_seq_c(pool, clusters, cycles_per_ms)
        else:
            c_src = _C_RENDERERS[mode](pool, clusters, res.row_latencies[pool],
                                       cycles_per_ms, n_iter)
        (test_dir / 'main.c').write_text(c_src)
        (test_dir / 'CMakeLists.txt').write_text(_render_cmake_txt())

        print(f'  [{pool}]  {test_dir / "main.c"}')
        print(f'  [{pool}]  {test_dir / "CMakeLists.txt"}')
        if gen_preload:
            elf_path = _build_preload_elf(test_dir, pool_safe, int(clusters[0].first_din))
            print(f'  [{pool}]  {elf_path}')


# ---------------------------------------------------------------------------
# Plotting
# ---------------------------------------------------------------------------

def _plot_cluster_split(plt, pool: str, clusters: List[ClusterSplit],
                        throughput_ms: float, overlap: bool, out_path: Path) -> None:
    """Bar plot of per-cluster wall time: compute (red) vs transfer (green).

    Serialized split -> stacked bars   (wall = compute + transfer).
    Overlapped split -> grouped bars    (wall = max(compute, transfer)).
    The y-axis is capped at throughput_ms so per-cluster headroom is visible.
    """
    x        = list(range(len(clusters)))
    ids      = [str(cs.idx)        for cs in clusters]
    compute  = [cs.latency         for cs in clusters]
    transfer = [cs.transfer_ms     for cs in clusters]

    fig, ax = plt.subplots(figsize=(max(4.0, 1.1 * len(clusters) + 2.0), 4.0))
    if overlap:
        w = 0.4
        ax.bar([xi - w / 2 for xi in x], compute,  width=w, color='red',   label='compute')
        ax.bar([xi + w / 2 for xi in x], transfer, width=w, color='green', label='transfer')
        tag = 'CWT'
    else:
        w = 0.6
        ax.bar(x, compute,  width=w, color='red',                   label='compute')
        ax.bar(x, transfer, width=w, bottom=compute, color='green', label='transfer')
        tag = 'CTT'

    ax.axhline(throughput_ms, color='black', linestyle='--', linewidth=1.0)
    ax.set_ylim(0.0, throughput_ms)          # fix the max at the throughput budget
    ax.set_xticks(x)
    ax.set_xticklabels(ids)
    ax.set_xlabel('cluster ID')
    ax.set_ylabel('wall time [ms]')
    ax.set_title(f'{pool} -- {tag}')
    ax.legend(loc='upper right', fontsize='small')
    fig.tight_layout()
    fig.savefig(out_path, dpi=120)
    plt.close(fig)


def generate_plots(res: Results, throughput_ms: float, mode: Optional[str],
                   out_dir: Path) -> None:
    """Save per-pool bar plots of the relevant cluster split(s) to out_dir.

    Plots the serialized split for seq/ctt, the overlapped split for cwt, and
    both when mode is None -- one PNG per (pool, split).  matplotlib is imported
    lazily; if it is missing the plots are skipped with a warning.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')   # headless: write PNGs, never open a window
        import matplotlib.pyplot as plt
    except ImportError:
        print('Warning: matplotlib not installed; --plot skipped '
              '(pip install matplotlib).')
        return

    want_serial  = mode in (None, 'seq', 'ctt')
    want_overlap = mode in (None, 'cwt')
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f'\nSaving cluster-split plots -> {out_dir}/')
    for p in res.clusters:
        psafe = p.replace(' ', '_')
        if want_serial and res.clusters.get(p):
            path = out_dir / f'CTT_split_{psafe}.png'
            _plot_cluster_split(plt, p, res.clusters[p], throughput_ms, False, path)
            print(f'  [plot] {path}')
        if want_overlap and res.clusters_overlap.get(p):
            path = out_dir / f'CWT_split_{psafe}.png'
            _plot_cluster_split(plt, p, res.clusters_overlap[p], throughput_ms, True, path)
            print(f'  [plot] {path}')


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def load_config(path: str) -> dict:
    with open(path) as f:
        return json.load(f)


def run(template_path: str, config_path: str) -> Results:
    """Programmatic entry point: returns the populated Results struct."""
    cfg  = load_config(config_path)
    grid = read_grid(template_path)

    header_row = find_header_row(grid)
    cols       = build_column_map(grid[header_row])
    data_start = header_row + 1
    data_end   = find_data_end(grid, cols, data_start)   # exclusive
    check_for_errors(grid, cols, data_start, data_end - 1)
    layers     = parse_layers(grid, cols, data_start, data_end - 1)
    if not layers:
        raise ValueError("No data rows found below the header row")

    return compute(layers, cfg), layers


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument('template', help='Path to template (.xlsx, .xlsm, .ods or .csv)')
    ap.add_argument('config',   help='Path to JSON HW config')
    ap.add_argument(
        '--mode', choices=['seq', 'ctt', 'cwt'], metavar='{seq,ctt,cwt}',
        help='Limit to one pipeline flavour: seq = sequential, '
             'ctt = compute-then-transfer, cwt = compute-while-transfer.  '
             "Prints only that mode's cluster split and generates only its "
             'test dir.  Omit to print both splits and generate all three.  '
             'Generation requires "throughput" in config.',
    )
    ap.add_argument('--iterations', type=int, default=1, metavar='N',
                    help='Items streamed through the pipeline for ctt/cwt '
                         '(default: 1; ignored by seq)')
    ap.add_argument('--out', metavar='DIR',
                    help='Output directory for the generated test (requires '
                         '--mode).  Without --mode each mode uses its default '
                         'dir (07_pipeline_seq / 07_pipeline_sim_ctt / '
                         '07_pipeline_sim_cwt under flex_cluster_sdk/tests/).')
    ap.add_argument('--gen-preload', action='store_true',
                    help='Build the HBM preload .elf (preload_<pool>.elf) directly '
                         '(requires the riscv32-unknown-elf-* toolchain).  By default '
                         'no preload is generated.')
    ap.add_argument('--plot', action='store_true',
                    help='Save a per-pool bar plot of the cluster split(s) as PNGs '
                         'under ./split_plots/: compute (red) and transfer (green), '
                         'stacked for the serialized split (seq/ctt) and side-by-side '
                         'for the overlapped split (cwt), y-axis capped at throughput.  '
                         'Which split(s) follow --mode (both when omitted).')
    args = ap.parse_args()

    if args.iterations < 1:
        ap.error('--iterations must be >= 1')
    if args.out and not args.mode:
        ap.error('--out requires --mode (without --mode all three modes are '
                 'generated to their default directories)')

    cfg = load_config(args.config)
    res, layers = run(args.template, args.config)

    # Always print the tables; the cluster split shows both flavours when no
    # --mode is given, or only the selected one otherwise.
    print_results(layers, res, args.mode)

    # Generate the selected mode, or all three when --mode is omitted.
    modes = [args.mode] if args.mode else ['seq', 'ctt', 'cwt']
    if not res.clusters:
        print('\nWarning: no cluster splits found -- nothing generated. '
              'Add "throughput" to config.json and re-run.')
    else:
        frequency     = cfg.get('frequency', 1000000000)
        cycles_per_ms = int(frequency // 1000)
        for m in modes:
            out = args.out if args.out else str(DEFAULT_OUT[m])
            generate_tests(res, out, cycles_per_ms, m, args.iterations, args.gen_preload)

    if args.plot:
        if not res.clusters:
            print('\nWarning: no cluster splits to plot. '
                  'Add "throughput" to config.json and re-run.')
        else:
            generate_plots(res, float(cfg['throughput']), args.mode,
                           Path.cwd() / 'split_plots')


if __name__ == '__main__':
    main()
